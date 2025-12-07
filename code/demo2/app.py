"""
Flask Backend Integration for Enhanced Student Scoring System
Integrates the enhanced essay analyzer with your existing frontend
"""
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from student_analyzer import StudentAnalyzer, StudentScore
import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = 'your-secret-key-here-change-in-production'

# Initialize the enhanced analyzer
analyzer = StudentAnalyzer()

# Store analysis results in memory (you can replace with database)
analysis_results = []

# Sample data for testing
SAMPLE_DATA = {
    'high': {
        'studentId': 'STU_2025_001',
        'country': 'India',
        'gpa': 3.8,
        'curriculum': 'IGCSE/IB',
        'travelHistory': 'Multiple listed',
        'essayText': """Throughout my academic journey, I have developed a profound passion for computer science and artificial intelligence. My experiences in leading the robotics team at my high school taught me the importance of collaboration, innovation, and perseverance. I successfully organized three international STEM workshops that brought together students from diverse backgrounds to explore cutting-edge technology.

I am particularly drawn to Washington State University because of its renowned research programs in machine learning and its commitment to fostering a diverse academic community. I have spent considerable time analyzing WSU's curriculum and believe that the interdisciplinary approach aligns perfectly with my goal of developing AI solutions for healthcare accessibility in underserved communities.

My vision extends beyond personal achievement. Having witnessed the healthcare challenges in rural areas of my country, I am determined to leverage technology to create meaningful change. I have already initiated a pilot project that uses basic machine learning algorithms to predict disease outbreaks in my local community, and I am eager to expand this work with the resources and mentorship available at WSU.

I recognize that success requires not just technical skills but also cultural awareness and adaptability. My experiences traveling to multiple countries and engaging with different educational systems have prepared me to thrive in WSU's international environment. I am committed to contributing to campus life through student organizations, research opportunities, and community service initiatives.""",
        'negFactors': []
    },
    'medium': {
        'studentId': 'STU_2025_002',
        'country': 'Vietnam',
        'gpa': 3.2,
        'curriculum': 'Standard Intl Secondary',
        'travelHistory': '1 listed or multiple non-listed',
        'essayText': """I want to study at Washington State University because it is a good school. I have always dreamed of studying in America since I was a child. My passion for learning drives me to pursue my goals.

I believe WSU will help me reach my full potential. I am a hard worker and I always give 110% effort. I want to make a difference in the world and give back to my community. This has been my dream for as long as I can remember.

I think I would be a good fit for WSU because I am dedicated and motivated. I am looking forward to the opportunities that await me. Thank you for considering my application.""",
        'negFactors': ['bankDocsPending']
    },
    'low': {
        'studentId': 'STU_2025_003',
        'country': 'Pakistan',
        'gpa': 2.4,
        'curriculum': 'N/A',
        'travelHistory': 'No travel abroad',
        'essayText': """I want study USA. Good university. I work hard and study. My family support me. I want learn computer. Thank you.""",
        'negFactors': ['reqAppFeeWaiver', 'cannotPayFee', 'earlyI20']
    }
}

@app.route('/')
def index():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('index.html', username=session.get('username', 'User'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # Accept JSON OR form data (frontend flexibility)
        data = request.get_json(silent=True)

        if data:
            username = data.get("username")
            password = data.get("password")
        else:
            # fallback for form submissions
            username = request.form.get("username")
            password = request.form.get("password")

        # Authentication
        if username == "admin" and password == "admin123":
            session["username"] = username
            return jsonify({"success": True, "redirect": url_for("index")})
        else:
            return jsonify({"success": False, "error": "Invalid credentials"})

    return render_template("login.html")

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/api/sample/<type>')
def get_sample(type):
    """Get sample student data"""
    if type not in SAMPLE_DATA:
        return jsonify({'error': 'Invalid sample type'}), 400
    
    return jsonify(SAMPLE_DATA[type])

@app.route('/api/analyze', methods=['POST'])
def analyze_student():
    """
    Main analysis endpoint - uses enhanced essay analyzer
    """
    try:
        data = request.get_json()
        
        # Extract data
        gpa = float(data.get('gpa', 0))
        curriculum = data.get('curriculum', '')
        travel_history = data.get('travelHistory', '')
        essay_text = data.get('essayText', '')
        neg_factors = data.get('negFactors', [])
        
        # Validate inputs
        if not curriculum or not travel_history:
            return jsonify({'error': 'Missing required fields'}), 400
        
        # Run enhanced analysis
        result = analyzer.analyze_student(
            gpa=gpa,
            curriculum=curriculum,
            travel_history=travel_history,
            essay_text=essay_text,
            neg_factors=neg_factors
        )
        
        # Store result for later export
        analysis_record = {
            'timestamp': datetime.now().isoformat(),
            'studentId': data.get('studentId', 'N/A'),
            'country': data.get('country', 'N/A'),
            'gpa': gpa,
            'curriculum': curriculum,
            'travelHistory': travel_history,
            'essayLength': len(essay_text),
            'negFactors': ', '.join(neg_factors),
            'posScore': result.pos_score,
            'negScore': result.neg_score,
            'finalScore': result.final_score,
            'rankEstimate': result.rank_estimate,
            'recommendation': result.recommendation,
            'clarityFocus': result.essay_analysis.clarity_focus,
            'developmentOrg': result.essay_analysis.development_organization,
            'creativityStyle': result.essay_analysis.creativity_style,
            'essayRubricScore': result.essay_analysis.rubric_score,
            'grammarScore': result.essay_analysis.grammar_score,
            'coherenceScore': result.essay_analysis.coherence_score,
            'vocabularyRichness': result.essay_analysis.vocabulary_richness,
            'analysisConfidence': result.overall_confidence
        }
        
        analysis_results.append(analysis_record)
        
        # Format response for frontend
        response = {
            'posScore': result.pos_score,
            'negScore': result.neg_score,
            'finalScore': result.final_score,
            'rankEstimate': result.rank_estimate,
            'recommendation': result.recommendation,
            'confidence': result.overall_confidence,
            'breakdown': result.breakdown,
            'essayAnalysis': {
                'clarityFocus': result.essay_analysis.clarity_focus,
                'developmentOrganization': result.essay_analysis.development_organization,
                'creativityStyle': result.essay_analysis.creativity_style,
                'totalScore': result.essay_analysis.total_score,
                'rubricScore': result.essay_analysis.rubric_score,
                'weightedScore': result.essay_analysis.weighted_score,
                'grammarScore': result.essay_analysis.grammar_score,
                'coherenceScore': result.essay_analysis.coherence_score,
                'vocabularyRichness': result.essay_analysis.vocabulary_richness,
                'insights': result.essay_analysis.insights,
                'strengths': result.essay_analysis.strengths,
                'weaknesses': result.essay_analysis.weaknesses,
                'confidence': result.essay_analysis.analysis_confidence
            }
        }
        
        return jsonify(response)
    
    except Exception as e:
        print(f"Analysis error: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/csv', methods=['POST'])
def export_csv():
    """Export analysis results as CSV"""
    try:
        output = io.StringIO()
        
        if not analysis_results:
            return jsonify({'error': 'No data to export'}), 400
        
        # Define CSV headers
        headers = [
            'Timestamp', 'Student ID', 'Country', 'GPA', 'Curriculum', 
            'Travel History', 'Essay Length', 'Neg Factors',
            'POS Score', 'NEG Score', 'Final Score', 'Rank Estimate',
            'Clarity & Focus', 'Development & Org', 'Creativity & Style',
            'Essay Rubric Score', 'Grammar Score', 'Coherence Score',
            'Vocabulary Richness', 'Analysis Confidence', 'Recommendation'
        ]
        
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        
        for record in analysis_results:
            writer.writerow({
                'Timestamp': record['timestamp'],
                'Student ID': record['studentId'],
                'Country': record['country'],
                'GPA': record['gpa'],
                'Curriculum': record['curriculum'],
                'Travel History': record['travelHistory'],
                'Essay Length': record['essayLength'],
                'Neg Factors': record['negFactors'],
                'POS Score': record['posScore'],
                'NEG Score': record['negScore'],
                'Final Score': record['finalScore'],
                'Rank Estimate': record['rankEstimate'],
                'Clarity & Focus': record['clarityFocus'],
                'Development & Org': record['developmentOrg'],
                'Creativity & Style': record['creativityStyle'],
                'Essay Rubric Score': record['essayRubricScore'],
                'Grammar Score': record['grammarScore'],
                'Coherence Score': record['coherenceScore'],
                'Vocabulary Richness': record['vocabularyRichness'],
                'Analysis Confidence': record['analysisConfidence'],
                'Recommendation': record['recommendation']
            })
        
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'student_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
    
    except Exception as e:
        print(f"CSV export error: {e}")
        return jsonify({'error': str(e)}), 500

# Excel export
@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    try:
        if not analysis_results:
            return jsonify({'error': 'No data to export'}), 400
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Analysis"
        headers = [
            'Timestamp', 'Student ID', 'Country', 'GPA', 'Curriculum', 
            'Travel History', 'Essay Length', 'Neg Factors',
            'POS Score', 'NEG Score', 'Final Score', 'Rank Estimate',
            'Clarity & Focus', 'Development & Org', 'Creativity & Style',
            'Essay Rubric Score', 'Grammar Score', 'Coherence Score',
            'Vocabulary Richness', 'Analysis Confidence', 'Recommendation'
        ]
        header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row_idx, record in enumerate(analysis_results, start=2):
            ws.cell(row=row_idx, column=1, value=record['timestamp'])
            ws.cell(row=row_idx, column=2, value=record['studentId'])
            ws.cell(row=row_idx, column=3, value=record['country'])
            ws.cell(row=row_idx, column=4, value=record['gpa'])
            ws.cell(row=row_idx, column=5, value=record['curriculum'])
            ws.cell(row=row_idx, column=6, value=record['travelHistory'])
            ws.cell(row=row_idx, column=7, value=record['essayLength'])
            ws.cell(row=row_idx, column=8, value=record['negFactors'])
            ws.cell(row=row_idx, column=9, value=record['posScore'])
            ws.cell(row=row_idx, column=10, value=record['negScore'])
            ws.cell(row=row_idx, column=11, value=record['finalScore'])
            ws.cell(row=row_idx, column=12, value=record['rankEstimate'])
            ws.cell(row=row_idx, column=13, value=record['clarityFocus'])
            ws.cell(row=row_idx, column=14, value=record['developmentOrg'])
            ws.cell(row=row_idx, column=15, value=record['creativityStyle'])
            ws.cell(row=row_idx, column=16, value=record['essayRubricScore'])
            ws.cell(row=row_idx, column=17, value=record['grammarScore'])
            ws.cell(row=row_idx, column=18, value=record['coherenceScore'])
            ws.cell(row=row_idx, column=19, value=record['vocabularyRichness'])
            ws.cell(row=row_idx, column=20, value=record['analysisConfidence'])
            ws.cell(row=row_idx, column=21, value=record['recommendation'])
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'student_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        print(f"Excel export error: {e}")
        return jsonify({'error': str(e)}), 500

# Analytics dashboard
@app.route('/analytics')
def analytics():
    if 'username' not in session:
        return redirect(url_for('login'))

    total_analyzed = len(analysis_results)
    distribution = {'highPotential': 0, 'mediumRisk': 0, 'highRisk': 0}
    country_counts = {}
    total_final_score = 0
    total_grammar_score = 0

    for r in analysis_results:
        final_score = r.get('finalScore', 0)
        total_final_score += final_score
        total_grammar_score += r.get('grammarScore', 0)

        recommendation = r.get('recommendation', '').lower()
        if 'highly recommended' in recommendation or final_score >= 30:
            distribution['highPotential'] += 1
        elif 'recommended' in recommendation or 20 <= final_score < 30:
            distribution['mediumRisk'] += 1
        else:
            distribution['highRisk'] += 1

        country = r.get('country', 'Unknown')
        country_counts[country] = country_counts.get(country, 0) + 1

    avg_final_score = round(total_final_score / total_analyzed, 2) if total_analyzed else 0
    avg_grammar_score = round(total_grammar_score / total_analyzed, 2) if total_analyzed else 0

    # Top countries by count
    top_countries = sorted(
        [{'country': k, 'count': v} for k, v in country_counts.items()],
        key=lambda x: x['count'],
        reverse=True
    )[:10]

    # Average score per country
    country_scores = {}
    for r in analysis_results:
        country = r.get('country', 'Unknown')
        country_scores.setdefault(country, []).append(r.get('finalScore', 0))
    avg_scores_per_country = {c: round(sum(scores)/len(scores), 2) for c, scores in country_scores.items()}

    analytics_data = {
        'totalStudents': total_analyzed,
        'avgFinalScore': avg_final_score,
        'avgGrammarScore': avg_grammar_score,
        'distribution': distribution,
        'countryStats': top_countries,
        'avgScoresPerCountry': avg_scores_per_country,
        'results': analysis_results[-10:]
    }

    return render_template('analytics.html', username=session.get('username'), analytics=analytics_data)

# API endpoint for frontend charts
@app.route('/api/analytics/summary')
def analytics_summary():
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401

    total_analyzed = len(analysis_results)
    distribution = {'highPotential': 0, 'mediumRisk': 0, 'highRisk': 0}
    country_counts = {}
    total_final_score = 0
    total_grammar_score = 0

    for r in analysis_results:
        final_score = r.get('finalScore', 0)
        total_final_score += final_score
        total_grammar_score += r.get('grammarScore', 0)

        recommendation = r.get('recommendation', '').lower()
        if 'highly recommended' in recommendation or final_score >= 30:
            distribution['highPotential'] += 1
        elif 'recommended' in recommendation or 20 <= final_score < 30:
            distribution['mediumRisk'] += 1
        else:
            distribution['highRisk'] += 1

        country = r.get('country', 'Unknown')
        country_counts[country] = country_counts.get(country, 0) + 1

    avg_final_score = round(total_final_score / total_analyzed, 2) if total_analyzed else 0
    avg_grammar_score = round(total_grammar_score / total_analyzed, 2) if total_analyzed else 0

    top_countries = sorted(
        [{'country': k, 'count': v} for k, v in country_counts.items()],
        key=lambda x: x['count'],
        reverse=True
    )[:10]

    country_scores = {}
    for r in analysis_results:
        country = r.get('country', 'Unknown')
        country_scores.setdefault(country, []).append(r.get('finalScore', 0))
    avg_scores_per_country = {c: round(sum(scores)/len(scores), 2) for c, scores in country_scores.items()}

    return jsonify({
        'totalStudents': total_analyzed,
        'avgFinalScore': avg_final_score,
        'avgGrammarScore': avg_grammar_score,
        'distribution': distribution,
        'countryStats': top_countries,
        'avgScoresPerCountry': avg_scores_per_country
    })

@app.route('/batch')
def batch():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('batch.html', username=session.get('username'))

@app.route('/financial')
def financial():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('financial.html', username=session.get('username'))

@app.route('/transcript')
def transcript():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('transcript.html', username=session.get('username'))

if __name__ == '__main__':
    app.run(debug=True, port=5000)