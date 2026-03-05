"""
Flask Backend Integration for Enhanced Student Scoring System
Integrates the enhanced essay analyzer with your existing frontend
"""
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from services.student_analyzer import StudentAnalyzerSafe
from db.database import get_db
from db.report_generator import generate_report as build_pdf_report
from services.nlp_service import NLPService
import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

nlp_service = NLPService()
app = Flask(__name__)
app.secret_key = 'your-secret-key-here-change-in-production'

# Initialize the enhanced analyzer
analyzer = StudentAnalyzerSafe()
analysis_results = []

# --- SAMPLE DATA ---
SAMPLE_DATA = {
    'high': {
        'studentId': 'STU_HIGH_001',
        'country': 'India',
        'gpa': 3.9,
        'curriculum': 'US HS/University',
        'travelHistory': 'SEVIS/Multiple US trips',
        'essayText': 'Throughout my academic journey, I have consistently demonstrated a passion for engineering and innovation. My research experience at IIT Delhi, combined with my internship at a leading tech firm, has prepared me well for graduate study. I am particularly interested in machine learning applications in healthcare, and I believe WSU\'s program aligns perfectly with my goals.',
        'negFactors': []
    },
    'medium': {
        'studentId': 'STU_MED_001',
        'country': 'Vietnam',
        'gpa': 3.2,
        'curriculum': 'IGCSE/IB',
        'travelHistory': 'Multiple listed',
        'essayText': 'I have always been passionate about computer science. During my undergraduate studies, I worked on several projects involving data analysis and software development. I hope to further my education at WSU and contribute to the research community.',
        'negFactors': ['bankDocsPending']
    },
    'low': {
        'studentId': 'STU_LOW_001',
        'country': 'Nigeria',
        'gpa': 2.4,
        'curriculum': 'Standard Intl Secondary',
        'travelHistory': 'No travel abroad',
        'essayText': 'I want to study at WSU because it is a good university. I will work hard and do my best.',
        'negFactors': ['reqAppFeeWaiver', 'cannotPayFee', 'bankDocsPending']
    }
}

# --- ROUTES ---
@app.route('/')
def index():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('index.html', username=session.get('username', 'User'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        data = request.get_json(silent=True) or request.form
        username = data.get("username")
        password = data.get("password")
        if not username or not password:
            return jsonify({"success": False, "error": "Username and password required"})
        db = get_db()
        user = db.authenticate_user(username, password)
        if user:
            session["user_id"] = user['id']
            session["username"] = user['username']
            session["full_name"] = user.get('full_name') or user['username']
            session["role"] = user['role']
            return jsonify({"success": True, "redirect": url_for("index")})
        else:
            return jsonify({"success": False, "error": "Invalid credentials"})
    return render_template("login.html")

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

def require_login():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return None

@app.route('/api/sample/<type>')
def get_sample(type):
    if type not in SAMPLE_DATA:
        return jsonify({'error': 'Invalid sample type'}), 400
    return jsonify(SAMPLE_DATA[type])

# --- STUDENT ANALYSIS ---
@app.route("/api/analyze", methods=["POST"])
def analyze_student():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No input data provided'}), 400

        essay = data.get("essayText") or data.get("essay")
        prompt = data.get("prompt", "")
        if not essay:
            return jsonify({'error': 'No essay text provided'}), 400

        sentiment = nlp_service.analyze_sentiment(essay)
        similarity = nlp_service.compute_similarity(essay, prompt)

        gpa = float(data.get('gpa', 0))
        curriculum = data.get('curriculum', '')
        travel_history = data.get('travelHistory', '')
        neg_factors = data.get('negFactors', [])

        result_resp = analyzer.analyze_student_safe(
            gpa=gpa,
            curriculum=curriculum,
            travel_history=travel_history,
            essay_text=essay,
            neg_factors=neg_factors
        )

        # Convert to dict safely
        if isinstance(result_resp, dict):
            result = result_resp
        elif hasattr(result_resp, "model_dump"):
            result = result_resp.model_dump()
        elif hasattr(result_resp, "dict") and callable(getattr(result_resp, "dict", None)):
            result = result_resp.dict()
        else:
            try:
                result = dict(result_resp)
            except:
                result = result_resp

        response = {
            'posScore': float(result.get('pos_score', 0)),
            'negScore': float(result.get('neg_score', 0)),
            'finalScore': float(result.get('final_score', 0)),
            'breakdown': result.get('breakdown', {}),
            'rankEstimate': result.get('rank_estimate', "N/A"),
            'recommendation': result.get('recommendation', "No recommendation"),
            'essayAnalysis': {
                'clarity_focus': result.get('essay_analysis', {}).get('clarity_focus', ''),
                'development_organization': result.get('essay_analysis', {}).get('development_organization', ''),
                'creativity_style': result.get('essay_analysis', {}).get('creativity_style', ''),
                'rubric_score': float(result.get('essay_analysis', {}).get('rubric_score', 0)),
                'grammar_score': float(result.get('essay_analysis', {}).get('grammar_score', 0)),
                'coherence_score': float(result.get('essay_analysis', {}).get('coherence_score', 0)),
                'vocabulary_richness': float(result.get('essay_analysis', {}).get('vocabulary_richness', 0)),
                'insights': result.get('essay_analysis', {}).get('insights', []),
            },
            'overall_confidence': float(result.get('overall_confidence', 0))
        }

        # ── Save to in-memory results so analytics dashboard is populated ──
        analysis_results.append({
            'timestamp': datetime.now().isoformat(),
            'studentId': data.get('studentId', 'N/A'),
            'country': data.get('country', 'N/A'),
            'gpa': gpa,
            'curriculum': curriculum,
            'travelHistory': travel_history,
            'essayLength': len(essay),
            'negFactors': ', '.join(neg_factors),
            'posScore': response['posScore'],
            'negScore': response['negScore'],
            'finalScore': response['finalScore'],
            'rankEstimate': response['rankEstimate'],
            'recommendation': response['recommendation'],
            'clarityFocus': response['essayAnalysis'].get('clarity_focus', ''),
            'developmentOrg': response['essayAnalysis'].get('development_organization', ''),
            'creativityStyle': response['essayAnalysis'].get('creativity_style', ''),
            'essayRubricScore': response['essayAnalysis'].get('rubric_score', 0),
            'grammarScore': response['essayAnalysis'].get('grammar_score', 0),
            'coherenceScore': response['essayAnalysis'].get('coherence_score', 0),
            'vocabularyRichness': response['essayAnalysis'].get('vocabulary_richness', 0),
            'analysisConfidence': response.get('overall_confidence', 0)
        })

        return jsonify(response)

    except Exception as e:
        print(f"Analysis error: {e}")
        return jsonify({'error': str(e)}), 500

# --- BATCH ANALYSIS ---
@app.route('/api/batch-analyze', methods=['POST'])
def batch_analyze():
    try:
        data = request.get_json()
        students = data.get('students', [])
        if not students:
            return jsonify({'error': 'No students provided'}), 400

        batch_results = []

        for student_data in students:
            try:
                student_id = student_data.get('studentId', 'N/A')
                country = student_data.get('country', 'N/A')
                gpa = float(student_data.get('gpa', 0))
                curriculum = student_data.get('curriculum', '')
                travel_history = student_data.get('travelHistory', '')
                essay_text = student_data.get('essayText', '')
                neg_factors = student_data.get('negFactors', [])

                if not curriculum or not travel_history:
                    batch_results.append({
                        'studentId': student_id,
                        'success': False,
                        'error': 'Missing required fields'
                    })
                    continue

                result_resp = analyzer.analyze_student_safe(
                    gpa=gpa,
                    curriculum=curriculum,
                    travel_history=travel_history,
                    essay_text=essay_text,
                    neg_factors=neg_factors
                )

                # Convert to dict safely
                if isinstance(result_resp, dict):
                    result = result_resp
                elif hasattr(result_resp, "model_dump"):
                    result = result_resp.model_dump()
                elif hasattr(result_resp, "dict") and callable(getattr(result_resp, "dict", None)):
                    result = result_resp.dict()
                else:
                    try:
                        result = dict(result_resp)
                    except:
                        result = result_resp

                ea = result.get('essay_analysis', {})

                analysis_record = {
                    'timestamp': datetime.now().isoformat(),
                    'studentId': student_id,
                    'country': country,
                    'gpa': gpa,
                    'curriculum': curriculum,
                    'travelHistory': travel_history,
                    'essayLength': len(essay_text),
                    'negFactors': ', '.join(neg_factors),
                    'posScore': float(result.get('pos_score', 0)),
                    'negScore': float(result.get('neg_score', 0)),
                    'finalScore': float(result.get('final_score', 0)),
                    'rankEstimate': result.get('rank_estimate', 'N/A'),
                    'recommendation': result.get('recommendation', 'No recommendation'),
                    'clarityFocus': ea.get('clarity_focus', ''),
                    'developmentOrg': ea.get('development_organization', ''),
                    'creativityStyle': ea.get('creativity_style', ''),
                    'essayRubricScore': float(ea.get('rubric_score', 0)),
                    'grammarScore': float(ea.get('grammar_score', 0)),
                    'coherenceScore': float(ea.get('coherence_score', 0)),
                    'vocabularyRichness': float(ea.get('vocabulary_richness', 0)),
                    'analysisConfidence': float(result.get('overall_confidence', 0))
                }

                analysis_results.append(analysis_record)

                batch_results.append({
                    'studentId': student_id,
                    'country': country,
                    'success': True,
                    'posScore': analysis_record['posScore'],
                    'negScore': analysis_record['negScore'],
                    'finalScore': analysis_record['finalScore'],
                    'rankEstimate': analysis_record['rankEstimate'],
                    'recommendation': analysis_record['recommendation']
                })

            except Exception as e:
                batch_results.append({
                    'studentId': student_data.get('studentId', 'Unknown'),
                    'success': False,
                    'error': str(e)
                })

        return jsonify({'results': batch_results})

    except Exception as e:
        print(f"Batch analysis error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/csv', methods=['POST'])
def export_csv():
    try:
        if not analysis_results:
            return jsonify({'error': 'No data to export'}), 400

        output = io.StringIO()
        headers = [
            'Timestamp', 'Student ID', 'Country', 'GPA', 'Curriculum',
            'Travel History', 'Essay Length', 'Neg Factors',
            'POS Score', 'NEG Score', 'Final Score', 'Rank Estimate',
            'Clarity & Focus', 'Development & Org', 'Creativity & Style',
            'Essay Rubric Score', 'Grammar Score', 'Coherence Score',
            'Vocabulary Richness', 'Analysis Confidence', 'Recommendation'
        ]

        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()

        for record in analysis_results:
            writer.writerow({
                'Timestamp': record['timestamp'],
                'Student ID': record['studentId'],
                'Country': record['country'],
                'GPA': record['gpa'],
                'Curriculum': record['curriculum'],
                'Travel History': record['travelHistory'],
                'Essay Length': record['essayLength'],
                'Neg Factors': record['negFactors'],
                'POS Score': record['posScore'],
                'NEG Score': record['negScore'],
                'Final Score': record['finalScore'],
                'Rank Estimate': record['rankEstimate'],
                'Clarity & Focus': record['clarityFocus'],
                'Development & Org': record['developmentOrg'],
                'Creativity & Style': record['creativityStyle'],
                'Essay Rubric Score': record['essayRubricScore'],
                'Grammar Score': record['grammarScore'],
                'Coherence Score': record['coherenceScore'],
                'Vocabulary Richness': record['vocabularyRichness'],
                'Analysis Confidence': record['analysisConfidence'],
                'Recommendation': record['recommendation']
            })

        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'student_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )

    except Exception as e:
        print(f"CSV export error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    try:
        if not analysis_results:
            return jsonify({'error': 'No data to export'}), 400

        wb = Workbook()
        ws = wb.active
        ws.title = "Student Analysis"

        headers = [
            'Timestamp', 'Student ID', 'Country', 'GPA', 'Curriculum',
            'Travel History', 'Essay Length', 'Neg Factors',
            'POS Score', 'NEG Score', 'Final Score', 'Rank Estimate',
            'Clarity & Focus', 'Development & Org', 'Creativity & Style',
            'Essay Rubric Score', 'Grammar Score', 'Coherence Score',
            'Vocabulary Richness', 'Analysis Confidence', 'Recommendation'
        ]

        header_fill = PatternFill(start_color='667eea', end_color='667eea', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_idx, record in enumerate(analysis_results, start=2):
            ws.cell(row=row_idx, column=1,  value=record['timestamp'])
            ws.cell(row=row_idx, column=2,  value=record['studentId'])
            ws.cell(row=row_idx, column=3,  value=record['country'])
            ws.cell(row=row_idx, column=4,  value=record['gpa'])
            ws.cell(row=row_idx, column=5,  value=record['curriculum'])
            ws.cell(row=row_idx, column=6,  value=record['travelHistory'])
            ws.cell(row=row_idx, column=7,  value=record['essayLength'])
            ws.cell(row=row_idx, column=8,  value=record['negFactors'])
            ws.cell(row=row_idx, column=9,  value=record['posScore'])
            ws.cell(row=row_idx, column=10, value=record['negScore'])
            ws.cell(row=row_idx, column=11, value=record['finalScore'])
            ws.cell(row=row_idx, column=12, value=record['rankEstimate'])
            ws.cell(row=row_idx, column=13, value=record['clarityFocus'])
            ws.cell(row=row_idx, column=14, value=record['developmentOrg'])
            ws.cell(row=row_idx, column=15, value=record['creativityStyle'])
            ws.cell(row=row_idx, column=16, value=record['essayRubricScore'])
            ws.cell(row=row_idx, column=17, value=record['grammarScore'])
            ws.cell(row=row_idx, column=18, value=record['coherenceScore'])
            ws.cell(row=row_idx, column=19, value=record['vocabularyRichness'])
            ws.cell(row=row_idx, column=20, value=record['analysisConfidence'])
            ws.cell(row=row_idx, column=21, value=record['recommendation'])

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'student_analysis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        print(f"Excel export error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/generate-report', methods=['POST'])
def generate_report_endpoint():
    """Generate report and save to database."""
    try:
        data           = request.get_json()
        student_data   = data.get('studentData', {})
        result_data    = data.get('resultData', {})
        staff_comments = data.get('staffComments', '')
        reviewer_name  = data.get('reviewerName', session.get('full_name', 'Staff'))
        fmt            = data.get('format', 'pdf').lower()
        analysis_id    = data.get('analysisId')

        student_id = student_data.get('studentId', 'N/A')
        db = get_db()

        if fmt == 'pdf':
            pdf_buffer = build_pdf_report(
                student_data   = student_data,
                result_data    = result_data,
                staff_comments = staff_comments,
                reviewer_name  = reviewer_name,
            )
            pdf_bytes = pdf_buffer.read()

            # Save to database using the actual signature: (analysis_id, student_id, user_id, pdf_blob)
            db.save_report(
                analysis_id = analysis_id,
                student_id  = student_id,
                user_id     = session.get('user_id', 0),
                pdf_blob    = pdf_bytes
            )

            return send_file(
                io.BytesIO(pdf_bytes),
                mimetype='application/pdf',
                as_attachment=True,
                download_name=f"admission_report_{student_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
            )

        else:  # txt
            now  = datetime.now().strftime('%B %d, %Y  %I:%M %p')
            sep  = '=' * 65
            thin = '-' * 65

            neg_str = ', '.join(student_data.get('negFactors', [])) or 'None'

            lines = [
                sep,
                'WSU OFFICE OF INTERNATIONAL PROGRAMS',
                'ADMISSION ANALYSIS REPORT — CONFIDENTIAL',
                sep,
                f"Student ID   : {student_id}",
                f"Country      : {student_data.get('country', 'N/A')}",
                f"Reviewed by  : {reviewer_name}",
                f"Report Date  : {now}",
                sep,
                '',
                'SCORE SUMMARY',
                thin,
                f"  POS Score    : +{float(result_data.get('posScore', 0)):.2f}",
                f"  NEG Score    :  -{abs(float(result_data.get('negScore', 0))):.2f}",
                f"  FINAL Score  :  {float(result_data.get('finalScore', 0)):.2f}",
                '',
                'RECOMMENDATION',
                thin,
                f"  {result_data.get('recommendation', 'N/A')}",
                '',
                'STAFF COMMENTS',
                thin,
                staff_comments.strip() if staff_comments.strip() else '(No additional comments recorded.)',
                '',
                sep,
            ]

            txt_content = '\n'.join(lines)

            # Save to database using the actual signature: (analysis_id, student_id, user_id, pdf_blob)
            db.save_report(
                analysis_id = analysis_id,
                student_id  = student_id,
                user_id     = session.get('user_id', 0),
                pdf_blob    = txt_content.encode('utf-8')
            )

            return send_file(
                io.BytesIO(txt_content.encode('utf-8')),
                mimetype='text/plain',
                as_attachment=True,
                download_name=f"admission_report_{student_id}_{datetime.now().strftime('%Y%m%d')}.txt"
            )

    except Exception as e:
        print(f"Report generation error: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/reports')
def reports_page():
    redirect_response = require_login()
    if redirect_response:
        return redirect_response
    db = get_db()
    reports = db.get_all_reports(limit=100) if hasattr(db, 'get_all_reports') else []
    return render_template('report.html',
                           username=session.get('full_name'),
                           reports=reports)


@app.route('/student/<student_id>')
def student_detail(student_id):
    redirect_response = require_login()
    if redirect_response:
        return redirect_response
    db = get_db()
    analyses = db.get_student_analyses(student_id) if hasattr(db, 'get_student_analyses') else []
    reports  = db.get_student_reports(student_id)  if hasattr(db, 'get_student_reports')  else []
    return render_template('student_detail.html',
                           username=session.get('full_name'),
                           student_id=student_id,
                           analyses=analyses,
                           reports=reports)


@app.route('/api/report/<int:report_id>/download')
def download_report(report_id):
    redirect_response = require_login()
    if redirect_response:
        return redirect_response
    db = get_db()
    report = db.get_report_by_id(report_id) if hasattr(db, 'get_report_by_id') else None
    if not report:
        return jsonify({'error': 'Report not found'}), 404
    return send_file(
        io.BytesIO(report['pdf_blob']),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f"report_{report['student_id']}_{report_id}.pdf"
    )


@app.route('/admin/users')
def admin_users():
    redirect_response = require_login()
    if redirect_response:
        return redirect_response
    if session.get('role') != 'admin':
        return "Access denied", 403
    db = get_db()
    users    = db.get_all_users()    if hasattr(db, 'get_all_users')    else []
    activity = db.get_activity_log(limit=50)
    return render_template('admin_users.html',
                           username=session.get('full_name'),
                           users=users,
                           activity=activity)


@app.route('/api/admin/create-user', methods=['POST'])
def create_user():
    if session.get('role') != 'admin':
        return jsonify({'error': 'Access denied'}), 403
    data      = request.get_json()
    username  = data.get('username')
    password  = data.get('password')
    full_name = data.get('full_name')
    email     = data.get('email')
    role      = data.get('role', 'reviewer')
    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400
    db      = get_db()
    user_id = db.create_user(username, password)
    if user_id:
        return jsonify({'success': True, 'user_id': user_id})
    else:
        return jsonify({'error': 'Username already exists'}), 400


@app.route('/dashboard')
def dashboard():
    redirect_response = require_login()
    if redirect_response:
        return redirect_response
    db = get_db()
    stats          = db.get_dashboard_stats()   if hasattr(db, 'get_dashboard_stats') else {}
    recent_reports = db.get_all_reports(limit=10) if hasattr(db, 'get_all_reports')   else []
    return render_template('dashboard.html',
                           username=session.get('full_name'),
                           stats=stats,
                           recent_reports=recent_reports)


def _build_analytics_data():
    """Shared helper used by both /analytics and /api/analytics/summary."""
    total_analyzed    = len(analysis_results)
    distribution      = {'highPotential': 0, 'mediumRisk': 0, 'highRisk': 0}
    country_counts    = {}
    total_final_score = 0
    total_grammar_score = 0

    for r in analysis_results:
        final_score = r.get('finalScore', 0)
        total_final_score   += final_score
        total_grammar_score += r.get('grammarScore', 0)

        recommendation = r.get('recommendation', '').lower()
        if 'highly recommended' in recommendation or final_score >= 7.5:
            distribution['highPotential'] += 1
        elif 'recommended' in recommendation or final_score >= 6.5:
            distribution['mediumRisk'] += 1
        else:
            distribution['highRisk'] += 1

        country = r.get('country', 'Unknown')
        country_counts[country] = country_counts.get(country, 0) + 1

    avg_final_score   = round(total_final_score   / total_analyzed, 2) if total_analyzed else 0
    avg_grammar_score = round(total_grammar_score / total_analyzed, 2) if total_analyzed else 0

    top_countries = sorted(
        [{'country': k, 'count': v} for k, v in country_counts.items()],
        key=lambda x: x['count'], reverse=True
    )[:10]

    country_scores = {}
    for r in analysis_results:
        country = r.get('country', 'Unknown')
        country_scores.setdefault(country, []).append(r.get('finalScore', 0))
    avg_scores_per_country = {
        c: round(sum(scores) / len(scores), 2) for c, scores in country_scores.items()
    }

    return {
        'totalStudents':       total_analyzed,
        'avgFinalScore':       avg_final_score,
        'avgGrammarScore':     avg_grammar_score,
        'distribution':        distribution,
        'countryStats':        top_countries,
        'avgScoresPerCountry': avg_scores_per_country,
        'results':             analysis_results[-10:]
    }


@app.route('/analytics')
def analytics():
    if 'username' not in session:
        return redirect(url_for('login'))
    analytics_data = _build_analytics_data()
    return render_template('analytics.html',
                           username=session.get('username'),
                           analytics=analytics_data)


@app.route('/api/analytics/summary')
def analytics_summary():
    if 'username' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    data = _build_analytics_data()
    data.pop('results', None)   # don't send full result list in summary API
    return jsonify(data)


@app.route('/batch')
def batch():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('batch.html', username=session.get('username'))


@app.route('/financial')
def financial():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('financial.html', username=session.get('username'))


@app.route('/transcript')
def transcript():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('transcript.html', username=session.get('username'))


if __name__ == '__main__':
    app.run(debug=True, port=5000)